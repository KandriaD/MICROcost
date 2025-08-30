import sys
import pandas as pd
import glob
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))
libraries_path = os.path.join(script_dir, "..", "Libraries")
if libraries_path not in sys.path:
    sys.path.append(libraries_path)

import bacteria
import media
import supplies
import antibiotics
import chemicals

special_forms = ["1_ml", "0.5_ml", "1_ul"]

# === Week Parsing Helper ===
def parse_week_string(week_str):
    """
    Convert a week string like 'Sept 2-3' or 'Sept30- Oct1' into a datetime (first day).
    Always returns a datetime (bad rows go to far future).
    """
    import datetime, re
    
    if not isinstance(week_str, str) or not week_str.strip():
        return datetime.datetime.max

    # Normalize month names
    week_str = week_str.replace("Sept", "Sep")

    # Ensure there's a space between month and day (e.g., "Sep30" → "Sep 30")
    week_str = re.sub(r'([A-Za-z]+)(\d)', r'\1 \2', week_str)

    try:
        first_part = week_str.split("-")[0].strip()
        return datetime.datetime.strptime(first_part + " 2025", "%b %d %Y")
    except Exception:
        return datetime.datetime.max


def load_course_info(course_df):
    return {
        "students": int(course_df.loc[0, "Students"]),
        "sections": int(course_df.loc[0, "Sections"]),
        "groups": int(course_df.loc[0, "Groups"]),
        "rooms": course_df.loc[0, "Rooms"].split(", ")
    }

def caluclate_total_qty(quantity, dist_type, course_info, form=None):
    multiplier = {
        "Per Student": course_info["students"],
        "Per Pair": course_info["students"] // 2,
        "Per Group": course_info["groups"],
        "Per Table": sum(5 if room.strip() == "113" else 6 for room in course_info['rooms']),
        "Per Section": course_info["sections"],
        "Per Room": len(course_info["rooms"]),
        "Per Course": 1
    }.get(dist_type, 1)

    if form in special_forms:
        return multiplier
    else:
        return quantity * multiplier
    
def get_all_weeks():
    folder_path = r"C:\Users\Kandriad\Desktop\ExpFiles"
    pattern = os.path.join(folder_path, "*Experiments.xlsx")
    experiment_files = glob.glob(pattern)

    all_weeks = set()
    for file in experiment_files:
        xls = pd.ExcelFile(file)
        schedule_df = xls.parse("ExperimentIndex")
        schedule_df.columns = schedule_df.columns.str.strip()
        weeks = schedule_df["Week"].unique()
        all_weeks.update(weeks)
    
    print(f"Found experiment files: {experiment_files}")
    print(f"Found weeks: {all_weeks}")

    # ✅ Sort weeks using parse_week_string
    return sorted(all_weeks, key=parse_week_string)


# == Lookup with synonym support ==
def lookup_by_name_or_key(name, library_dict, lib_label="item"):
    """Look up an item in a dictionary-based library (with optional synonyms)."""
    if not name or str(name).lower() == "nan":
        return None
    name_lower = str(name).strip().lower()

    for key, info in library_dict.items():
        # Match dictionary key
        if key.strip().lower() == name_lower:
            return info

        # Match "name" field
        if str(info.get("name", "")).strip().lower() == name_lower:
            return info

        # Match synonyms list
        synonyms = info.get("synonyms", [])
        if any(syn.strip().lower() == name_lower for syn in synonyms):
            return info

    print(f"Warning: '{name}' not found in {lib_label} library.")
    return None


# === Category validators ===
def is_valid_bacteria(name):
    # Exact key match only, no synonyms
    return name.strip() in bacteria.bacteria_list

def is_valid_media(name):
    return lookup_by_name_or_key(name, media.media_list, "media") is not None

def is_valid_supply(name):
    return lookup_by_name_or_key(name, supplies.supplies, "supply") is not None

def is_valid_chemical(name):
    return lookup_by_name_or_key(name, chemicals.chemical_list, "chemical") is not None

def is_valid_antibiotic(name):
    return lookup_by_name_or_key(name, antibiotics.antibiotics, "antibiotic") is not None


def validate_item(name, category):
    if category == "Biological":
        return is_valid_bacteria(name)
    elif category == "Uninoculated Media":
        return is_valid_media(name)
    elif category == "Chemical":
        return is_valid_chemical(name)
    elif category == "Supply":
        return is_valid_supply(name)
    elif category == "Antibiotics":
        if not is_valid_antibiotic(name):
            print(f"Warning: Antibiotic '{name}' not found in antibiotics list.")
            return False
        return True
    return False


# === Main Function to Gather Weekly Needs ===
def gather_weekly_needs(week_number=None, category_filter=None):

    folder_path = r"C:\Users\Kandriad\Desktop\ExpFiles"
    pattern = os.path.join(folder_path, "*Experiments.xlsx")
    experiment_files = glob.glob(pattern)

    weekly_biologicals = []
    weekly_supplies = []

    for file in experiment_files:
        course_name = os.path.basename(file).split("Experiments.xlsx")[0]
        xls = pd.ExcelFile(file)
        course_df = xls.parse("CourseInfo")
        course_info = load_course_info(course_df)

        schedule_df = xls.parse("ExperimentIndex")
        schedule_df.columns = schedule_df.columns.str.strip()

        week_exps = schedule_df[schedule_df["Week"] == week_number]

        for _, row in week_exps.iterrows():
            exp_title = row["Experiment Title"]
            sheet_name = str(row["SheetName"])
            if sheet_name.endswith('.0'):
                sheet_name = sheet_name[:-2]
            
            if sheet_name not in xls.sheet_names:
                print(f"Sheet '{sheet_name}' not found in file '{file}' — skipping.")
                continue
            
            experiment_df = xls.parse(sheet_name)
            experiment_df.columns = experiment_df.columns.str.strip()

            category_records = []

            for _, item in experiment_df.iterrows():
                item_category = str(item.get("Category", "")).strip()
                if not item_category or item_category.lower() == "nan":
                    continue
                item_name = item.get("Item/Organism")
                if pd.isna(item_name) and item_category == "Uninoculated Media":
                    item_name = item.get("Media Used")
                name = str(item_name).strip() if not pd.isna(item_name) else ""
                media_used = item.get("Media Used", None)
                form = item.get("Form", None)
                quantity = item.get("Quantity", 0)
                dist_type = str(item.get("DistributionType", "")).strip().title()

                try:
                    total_qty = caluclate_total_qty(quantity, dist_type, course_info, form)
                except Exception as e:
                    print(f"Error calculating quantity for {course_name} {exp_title} - {item_category} {name} {media_used}: {e}")
                    continue

                include = (
                    (category_filter is None) or
                    (category_filter == item_category) or
                    (category_filter == "Supplies" and item_category != "Biological")
                )

                if include:
                    record = {
                        "Week": week_number,
                        "Course": course_name,
                        "Category": item_category,
                        "Item/Organism": name,
                        "Media Used": media_used,
                        "Form": form,
                        "Total Quantity": total_qty,
                        "Notes": item.get("Notes", "")
                    }
                    if form in special_forms:
                        record["Volume"] = quantity
                                    
                    if item_category == "Biological":
                        if name in bacteria.bacteria_list:
                            weekly_biologicals.append(record)
                        else:
                            print(f"Warning: {name} not found in bacteria list for {course_name} {exp_title}")
                    elif category_filter == "Supplies":
                        weekly_supplies.append(record)  # Everything except Bacteria
                    category_records.append(record)

    if category_filter == "Biological":
        return weekly_biologicals
    elif category_filter == "Supplies":
        return weekly_supplies
    else:
        return weekly_biologicals, weekly_supplies
    

## exporting formatting
def style_and_format_excel(filename):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        consolas_font = Font(name='Consolas', size=11)
        italic_font = Font(name='Consolas', size=11, italic=True)

        headers = [str(cell.value).strip() for cell in ws[1]]
        try:
            item_col = headers.index("Item/Organism") + 1
        except ValueError:
            item_col = None
        try:
            category_col = headers.index("Category") + 1
        except ValueError:
            category_col = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            category = row[category_col - 1].value if category_col else None
            for cell in row:
                cell.font = consolas_font
            if item_col and ws.title == "Biologicals":
                item_value = row[item_col - 1].value
                if item_value and str(item_value).strip():
                    row[item_col - 1].font = italic_font

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(filename)


def export_all_weeks_to_two_sheets(filename):
    all_weeks = get_all_weeks()
    print(f"Exporting weekly needs for {len(all_weeks)} weeks...")
    all_biologicals = []
    all_supplies = []

    for week in all_weeks:  # already sorted with parse_week_string
        biologicals = gather_weekly_needs(week, "Biological")
        supplies = gather_weekly_needs(week, "Supplies")

        all_biologicals.extend(biologicals)
        all_supplies.extend(supplies)

    df_bio = pd.DataFrame(all_biologicals)
    if "Category" in df_bio.columns:
        df_bio = df_bio.drop(columns=["Category"])
    df_sup = pd.DataFrame(all_supplies)

    # Write both sheets in ONE ExcelWriter session
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_bio.to_excel(writer, sheet_name="Biologicals", index=False)
        df_sup.to_excel(writer, sheet_name="Supplies", index=False)

    # After writing is done and file is saved, apply styling
    style_and_format_excel(filename)


export_all_weeks_to_two_sheets("Fall25 Needs.xlsx")
